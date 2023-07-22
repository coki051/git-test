# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""
------------------------------------------
    Description :
    File Name : demo
    Author : Dade
    Create Time : 2022/4/21 9:37
------------------------------------------
    Change Activity:
        Modifier: 
        Modify time: 2022/4/21 9:37
------------------------------------------
"""

import os
import openpyxl
import xmindparser


class CasesBuild:

    def __init__(self):
        self.xmind = xmindparser
        self.name = input('请输入版本号：\n')
        self.mind_path = r'D:\Files\test_case_library\\'
        self.cases_path = r'D:\Files\test_case_library\历史用例\\'

    def get_basic(self):
        mind_files = os.listdir(self.mind_path)
        try:
            mind = self.mind_path + [name for name in mind_files if self.name in name][0]   # 通过版本号选取对应文件
            content = xmindparser.xmind_to_dict(mind)[0]['topic']                           # 获取xmind文件内容
            excel_name = content['title']
            case_info = content['topics']
            return excel_name, case_info
        except IndexError:
            raise RuntimeError('输入的版本号有误或xmind目录下不存在该文件...')

    def get_case(self, infos, cases, temp=''):
        for info in infos:
            info['title'] = '' if info['title'] is None else info['title']  # 如果为None 赋值为空 否则不变
            case = temp + info['title'] + '_'           # 递归组装各层级下的title  用_做间隔区分
            if 'topics' not in info:
                cases.append(case)
                continue
            branch = info['topics']
            self.get_case(branch, cases, temp=case)     # 开始递归
        return cases

    def create_xls(self):
        li = []
        name, infos = self.get_basic()
        cases = self.get_case(infos, li)
        cases_path = self.cases_path + name + '.xlsx'
        wk = openpyxl.Workbook()
        sheet_name = '测试用例'
        sheet = wk.active
        sheet.title = sheet_name
        sheet.append(['用例目录', '前置条件', '用例名称', '用例步骤', '预期结果', '是否冒烟测试'])
        wk.save(cases_path)
        for case in cases:
            file = openpyxl.load_workbook(cases_path)
            wk = file[file.sheetnames[0]]
            temp = case.split('_')[:-1]
            temp.insert(3, '') if len(temp) == 7 else None      # 如果长度只有7 默认没有前置条件 需要给空防出错
            row = wk.max_row                                    # 获取最大行数
            for i in range(2, len(temp)):                       # 用例导入目录从2开始
                value = temp[0] + '-' + temp[i-1] if i == 2 else temp[i]
                wk.cell(row+1, i-1, value)
                file.save(cases_path)


if __name__ == '__main__':
    CasesBuild().create_xls()
